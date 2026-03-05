import argparse
import json
import os
import random
import time
import uuid
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm


# =========================
# CONFIG
# =========================

WB_DESTS: List[int] = [-1029256, -1257786, -59202, -447422]

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0 Safari/537.36"
)

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
})

SEARCH_URL = "https://search.wb.ru/exactmatch/ru/common/v4/search"
DETAIL_URL = "https://card.wb.ru/cards/detail"


# =========================
# HELPERS
# =========================

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def gen_queryid() -> str:
    return uuid.uuid4().hex


def safe_get(d: Any, path: List[str], default=None):
    cur = d
    for p in path:
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return default
    return cur


def money_from_u(value_u: Optional[int]) -> Optional[float]:
    if value_u is None:
        return None
    return round(value_u / 100.0, 2)


def request_with_backoff(method: str, url: str, *, max_tries: int = 7, timeout: int = 25, **kwargs) -> requests.Response:
    last_exc: Optional[Exception] = None

    for attempt in range(1, max_tries + 1):
        time.sleep(random.uniform(0.35, 0.9))

        try:
            r = SESSION.request(method, url, timeout=timeout, **kwargs)

            if r.status_code in (429, 500, 502, 503, 504):
                cool = (1.6 ** attempt) + random.uniform(0.2, 1.5)
                time.sleep(cool)
                continue

            return r

        except requests.RequestException as e:
            last_exc = e
            cool = (1.6 ** attempt) + random.uniform(0.2, 1.5)
            time.sleep(cool)

    if last_exc:
        raise last_exc
    raise RuntimeError("request_with_backoff: exceeded retries")


def extract_products(search_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not isinstance(search_json, dict):
        return []

    p = search_json.get("products")
    if isinstance(p, list):
        return [x for x in p if isinstance(x, dict)]

    d = search_json.get("data")
    if isinstance(d, dict) and isinstance(d.get("products"), list):
        return [x for x in d["products"] if isinstance(x, dict)]

    return []


# =========================
# DATA MODEL
# =========================

@dataclass
class ProductRow:
    product_url: str
    article: int
    name: str
    price: Optional[float]
    description: Optional[str]
    image_urls: str
    characteristics_json: str
    seller_name: Optional[str]
    seller_url: Optional[str]
    sizes: str
    stock_total: Optional[int]
    rating: Optional[float]
    reviews_count: Optional[int]
    country_of_origin: Optional[str]


# =========================
# WB API
# =========================

def wb_search(query: str, page: int, dest: int) -> Dict[str, Any]:
    params = {
        "appType": 1,
        "curr": "rub",
        "dest": dest,
        "lang": "ru",
        "page": page,
        "query": query,
        "resultset": "catalog",
        "sort": "popular",
        "spp": 30,
        "suppressSpellcheck": False,
    }
    headers = {
        "x-queryid": gen_queryid(),
        "Referer": f"https://www.wildberries.ru/catalog/0/search.aspx?sort=popular&search={requests.utils.quote(query)}",
    }

    r = request_with_backoff("GET", SEARCH_URL, params=params, headers=headers)
    r.raise_for_status()
    return r.json()


def wb_detail(nm_ids: List[int], dest: int) -> Dict[str, Any]:
    params = {
        "appType": 1,
        "curr": "rub",
        "dest": dest,
        "nm": ",".join(map(str, nm_ids)),
    }
    headers = {"x-queryid": gen_queryid()}
    r = request_with_backoff("GET", DETAIL_URL, params=params, headers=headers)
    r.raise_for_status()
    return r.json()


def build_product_url(nm_id: int) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"


def build_seller_url(supplier_id: Optional[int]) -> Optional[str]:
    if not supplier_id:
        return None
    return f"https://www.wildberries.ru/seller/{supplier_id}"


# =========================
# DETAIL PARSING
# =========================

def parse_characteristics(prod: Dict[str, Any]) -> str:
    options = prod.get("options")
    if not isinstance(options, list):
        options = []
    norm = []
    for opt in options:
        if isinstance(opt, dict):
            norm.append({"name": opt.get("name"), "value": opt.get("value")})
    return json.dumps(norm, ensure_ascii=False, indent=2)


def parse_sizes_and_stock(prod: Dict[str, Any]) -> Tuple[str, Optional[int]]:
    sizes_list: List[str] = []
    stock_total = 0
    any_stock = False

    sizes = prod.get("sizes")
    if not isinstance(sizes, list):
        return "", None

    for s in sizes:
        if not isinstance(s, dict):
            continue

        size_name = s.get("name") or s.get("origName") or s.get("techSize")
        if size_name:
            sn = str(size_name)
            if sn not in sizes_list:
                sizes_list.append(sn)

        if isinstance(s.get("stocks"), list):
            for st in s["stocks"]:
                if isinstance(st, dict) and isinstance(st.get("qty"), int):
                    stock_total += st["qty"]
                    any_stock = True
        elif isinstance(s.get("qty"), int):
            stock_total += s["qty"]
            any_stock = True

    return ", ".join(sizes_list), (stock_total if any_stock else None)


def parse_country_of_origin(prod: Dict[str, Any]) -> Optional[str]:
    options = prod.get("options")
    if not isinstance(options, list):
        return None
    keys = {"страна производства", "страна-изготовитель", "страна бренда", "страна"}
    for opt in options:
        if not isinstance(opt, dict):
            continue
        name = str(opt.get("name", "")).strip().lower()
        val = opt.get("value")
        if name in keys and isinstance(val, str) and val.strip():
            return val.strip()
    return None


# =========================
# IMAGES (быстро: из search, если есть)
# =========================

def image_urls_from_search(p: Dict[str, Any]) -> str:
    """
    В search иногда есть ready-urls/ids.
    """
    # Часто search даёт только количество pics.
    # В detail у товара тоже есть pics (count), но url всё равно надо собирать через basket.
    # Чтобы не усложнять и не ловить блок, оставим пусто.
    return ""


# =========================
# EXCEL
# =========================

COLUMNS = [
    ("Ссылка на товар", "product_url"),
    ("Артикул", "article"),
    ("Название", "name"),
    ("Цена", "price"),
    ("Описание", "description"),
    ("Ссылки на изображения", "image_urls"),
    ("Характеристики (JSON)", "characteristics_json"),
    ("Название селлера", "seller_name"),
    ("Ссылка на селлера", "seller_url"),
    ("Размеры", "sizes"),
    ("Остатки (число)", "stock_total"),
    ("Рейтинг", "rating"),
    ("Количество отзывов", "reviews_count"),
    ("Страна производства", "country_of_origin"),
]


def autosize(ws):
    for i in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28


def save_xlsx(path: str, rows: List[ProductRow], sheet_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for i, (title, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=i, value=title)

    for r_i, row in enumerate(rows, start=2):
        for c_i, (_, attr) in enumerate(COLUMNS, start=1):
            ws.cell(row=r_i, column=c_i, value=getattr(row, attr))

    autosize(ws)
    wb.save(path)


# =========================
# MAIN PIPELINE
# =========================

def chunked(lst: List[int], n: int) -> List[List[int]]:
    return [lst[i:i + n] for i in range(0, len(lst), n)]


def pick_dest_with_results(query: str) -> Optional[int]:
    for d in WB_DESTS:
        js = wb_search(query, page=1, dest=d)
        prods = extract_products(js)
        total = js.get("total")
        print(f"[DEBUG] dest={d} status=200 total={total} products={len(prods)}")
        if prods:
            return d
    return None


def run(query: str, pages: int, out_dir: str):
    ensure_dir(out_dir)

    print(f"Поиск WB: '{query}'")
    chosen_dest = pick_dest_with_results(query)
    if chosen_dest is None:
        print("Варианты: WB режет сеть/регион или временно отдаёт пусто.")
        return

    print(f"Используем dest = {chosen_dest}")

    nm_ids: List[int] = []
    search_meta: Dict[int, Dict[str, Any]] = {}

    for page in range(1, pages + 1):
        js = wb_search(query, page=page, dest=chosen_dest)
        prods = extract_products(js)
        if not prods:
            print(f"Страница {page}: товаров нет.")
            break

        for p in prods:
            nm = p.get("id") or p.get("nmId")
            if isinstance(nm, int) and nm not in search_meta:
                search_meta[nm] = p
                nm_ids.append(nm)

        print(f"Страница {page}: собрано nmId = {len(nm_ids)}")

    if not nm_ids:
        print("Ничего не собрано.")
        return

    rows: List[ProductRow] = []

    for batch in tqdm(chunked(nm_ids, 50), desc="Карточки detail"):
        try:
            detail_js = wb_detail(batch, dest=chosen_dest)
            detail_products = safe_get(detail_js, ["data", "products"], default=[])
            if not isinstance(detail_products, list):
                detail_products = []
        except Exception:
            detail_products = []

        detail_by_id: Dict[int, Dict[str, Any]] = {}
        for dp in detail_products:
            if isinstance(dp, dict):
                did = dp.get("id") or dp.get("nmId")
                if isinstance(did, int):
                    detail_by_id[did] = dp

        for nm_id in batch:
            meta = search_meta.get(nm_id, {})
            prod = detail_by_id.get(nm_id, {})

            name = prod.get("name") or meta.get("name") or ""
            price = money_from_u(
                prod.get("salePriceU")
                or meta.get("salePriceU")
                or prod.get("priceU")
                or meta.get("priceU")
            )
            rating = prod.get("rating") or meta.get("rating")
            rating = float(rating) if isinstance(rating, (int, float)) else None

            feedbacks = prod.get("feedbacks") or meta.get("feedbacks")
            reviews_count = int(feedbacks) if isinstance(feedbacks, int) else None

            description = prod.get("description") or meta.get("description")

            supplier_id = prod.get("supplierId") or meta.get("supplierId")
            seller_name = prod.get("supplier") or meta.get("supplier")

            sizes_str, stock_total = parse_sizes_and_stock(prod) if prod else ("", None)
            characteristics_json = parse_characteristics(prod) if prod else "[]"
            country = parse_country_of_origin(prod) if prod else None

            image_urls = image_urls_from_search(meta)

            rows.append(
                ProductRow(
                    product_url=build_product_url(nm_id),
                    article=nm_id,
                    name=str(name),
                    price=price,
                    description=description,
                    image_urls=image_urls,
                    characteristics_json=characteristics_json,
                    seller_name=seller_name,
                    seller_url=build_seller_url(supplier_id if isinstance(supplier_id, int) else None),
                    sizes=sizes_str,
                    stock_total=stock_total,
                    rating=rating,
                    reviews_count=reviews_count,
                    country_of_origin=country,
                )
            )

    full_path = os.path.join(out_dir, "full_catalog.xlsx")
    save_xlsx(full_path, rows, "full")
    print(f"Полный каталог: {full_path} (строк: {len(rows)})")

    filtered = [
        r for r in rows
        if r.rating is not None
        and r.price is not None
        and r.country_of_origin is not None
        and r.rating >= 4.5
        and r.price <= 10000
        and r.country_of_origin.strip().lower() == "россия"
    ]

    filt_path = os.path.join(out_dir, "filtered_catalog.xlsx")
    save_xlsx(filt_path, filtered, "filtered")
    print(f"Фильтрованный каталог: {filt_path} (строк: {len(filtered)})")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--query", default="пальто из натуральной шерсти")
    ap.add_argument("--pages", type=int, default=10)
    ap.add_argument("--out", default="output")
    args = ap.parse_args()

    run(args.query, args.pages, args.out)


if __name__ == "__main__":
    main()